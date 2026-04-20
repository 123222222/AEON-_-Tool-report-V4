"""
services/excel.py
Download file Excel tu OneDrive, append 1 hang moi, upload lai.
"""
import os
import sys
import datetime
import tempfile
import requests
import base64
import re
import openpyxl

from auth import graph_session

EXCEL_SHARE_LINK = (
    "https://aeondelight-my.sharepoint.com/:x:/g/personal/dung_ho_aeondelight_biz"
    "/IQC4QAqLtp-mRqWCS_v596O8AXny2kJRSp6KVHCbb8knAb0?e=stWUfd"
)

CHART_EXCEL_SHARE_LINK = (
    "https://aeondelight-my.sharepoint.com/:x:/g/personal/phuc_nguyen_aeondelight_biz"
    "/IQDuEpUv8wt0T70HClrX1zROAeitk0A9LPQb-2YSc2fm9CA?e=N35nOA"
)

_excel_cache: dict = {"drive_id": None, "item_id": None}
_chart_excel_cache: dict = {"drive_id": None, "item_id": None}


def _log(msg):
    print(msg, flush=True)
    sys.stdout.flush()


def _resolve_excel_item(share_link: str = EXCEL_SHARE_LINK, cache: dict = None):
    if cache is None:
        cache = _excel_cache

    if cache["drive_id"] and cache["item_id"]:
        return cache["drive_id"], cache["item_id"]

    token   = graph_session.ensure_token()
    headers = {"Authorization": f"Bearer {token}"}
    encoded = base64.b64encode(share_link.encode("utf-8")).decode("utf-8")
    encoded = encoded.rstrip("=").replace("/", "_").replace("+", "-")
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/shares/u!{encoded}/driveItem",
        headers=headers
    )
    if r.status_code != 200:
        raise Exception(f"Khong resolve duoc file Excel: {r.status_code} {r.text[:200]}")
    data = r.json()
    cache["drive_id"] = data["parentReference"]["driveId"]
    cache["item_id"]  = data["id"]
    return cache["drive_id"], cache["item_id"]


def _to_int(value):
    try:
        if value is None or value == "":
            return 0
        return int(float(value))
    except Exception:
        return 0


def _download_excel(drive_id: str, item_id: str) -> str:
    token   = graph_session.ensure_token()
    headers = {"Authorization": f"Bearer {token}"}
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}",
        headers=headers
    )
    if r.status_code != 200:
        raise Exception(f"Khong lay duoc metadata Excel: {r.status_code}")
    download_url = r.json().get("@microsoft.graph.downloadUrl")
    if not download_url:
        raise Exception("Khong co downloadUrl cho file Excel")
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    r2  = requests.get(download_url, stream=True)
    for chunk in r2.iter_content(8192):
        tmp.write(chunk)
    tmp.close()
    return tmp.name


def _download_public_sharepoint_excel(share_link: str) -> str:
    """
    Download workbook directly from a public SharePoint sharing page.
    This avoids Graph auth for read-only chart rendering.
    """
    page = requests.get(share_link, timeout=30)
    if page.status_code != 200:
        raise Exception(f"Khong mo duoc share page: {page.status_code}")

    match = re.search(r'"FileGetUrl":"([^"]+)"', page.text)
    if not match:
        raise Exception("Khong tim thay FileGetUrl trong share page")

    file_url = match.group(1).replace('\\u0026', '&')
    download = requests.get(file_url, timeout=60)
    if download.status_code != 200:
        raise Exception(f"Khong tai duoc file Excel: {download.status_code}")

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.write(download.content)
    tmp.close()
    return tmp.name


def _upload_excel(drive_id: str, item_id: str, local_path: str):
    token   = graph_session.ensure_token()
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type":  "application/octet-stream",
    }
    with open(local_path, "rb") as f:
        data = f.read()
    import time as _time
    for attempt in range(3):
        r = requests.put(
            f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content",
            headers=headers, data=data
        )
        if r.status_code in (200, 201):
            return
        if r.status_code == 423:
            _log(f"[EXCEL] File bi lock, thu lai lan {attempt+1}/3 sau 5 giay...")
            _time.sleep(5)
            continue
        raise Exception(f"Upload Excel that bai: {r.status_code} {r.text[:200]}")
    raise Exception("Upload that bai sau 3 lan: file van bi lock. Hay dong file Excel tren OneDrive.")


def _normalize_time(t: str) -> str:
    if not t:
        return ""
    t = t.strip()
    if "AM" in t.upper() or "PM" in t.upper():
        for fmt in ("%I:%M %p", "%I:%M:%S %p"):
            try:
                return datetime.datetime.strptime(t.upper(), fmt.upper()).strftime("%H:%M")
            except ValueError:
                pass
    parts = t.split(":")
    try:
        return f"{int(parts[0]):02d}:{parts[1][:2]}"
    except Exception:
        return t


def _calc_downtime(start_time: str, start_date: str,
                   end_time: str, end_date: str):
    """
    Tinh tong phut tu start -> end, co tinh den ngay.
    - Neu qua ngay (overnight): overnight_min co gia tri, day/night_min = None
    - Neu cung ngay: tinh theo daytime/nighttime nhu cu
    Tra ve: (time_type, day_min, night_min, overnight_min)
    """
    if not start_time or not end_time:
        return "", None, None, None
    try:
        s_str = _normalize_time(start_time)
        e_str = _normalize_time(end_time)

        # Parse datetime day neu co
        fmt_dt = "%Y-%m-%d %H:%M"
        fmt_t  = "%H:%M"

        if start_date and end_date:
            s_dt = datetime.datetime.strptime(f"{start_date} {s_str}", fmt_dt)
            e_dt = datetime.datetime.strptime(f"{end_date} {e_str}", fmt_dt)
        else:
            s_dt = datetime.datetime.strptime(s_str, fmt_t)
            e_dt = datetime.datetime.strptime(e_str, fmt_t)

        # Neu end <= start va cung ngay -> thuc ra qua nua dem, cong 1 ngay cho end
        if e_dt <= s_dt:
            e_dt += datetime.timedelta(days=1)

        total_min = round((e_dt - s_dt).total_seconds() / 60, 2)

        # Overnight: tong thoi gian > 1440 phut (> 1 ngay)
        if total_min > 1440:
            _log(f"[EXCEL] -> Overnight, {total_min} phút")
            return "Overnight", None, None, total_min

        # Daytime: start tu 6:00 den 21:59
        # Nighttime: start tu 22:00 den 5:59
        if 6 <= s_dt.hour < 22:
            _log(f"[EXCEL] -> Daytime, {total_min} phút")
            return "Daytime", total_min, None, None
        else:
            _log(f"[EXCEL] -> Nighttime, {total_min} phút")
            return "Nighttime", None, total_min, None

    except Exception as ex:
        _log(f"[EXCEL] _calc_downtime error: {ex}")
        return "", None, None, None


def _get_next_no(ws) -> int:
    max_no = 0
    for row in range(9, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        try:
            n = int(val)
            if n > max_no:
                max_no = n
        except (TypeError, ValueError):
            pass
    return max_no + 1


def append_status_to_excel(
    site_name:   str,
    device:      str,
    pic:         str,
    alarm_type:  str = "",
    alarm_level: str = "",
    reason:      str = "",
    start_time:  str = "",
    start_date:  str = "",
    end_time:    str = "",
    end_date:    str = "",
    status:      str = "",
    description: str = "",
    processing:  str = "",
    week:        str = "",
) -> dict:
    """
    Format cot:
      A  = STT
      B  = Ten bo phan
      C  = Week
      D  = Ngay bat dau (dd/mm/yyyy)
      E  = Ngay ket thuc (dd/mm/yyyy)
      F  = rong
      G  = Nguoi phu trach
      H  = Ten thiet bi
      I  = Mo ta (Reason)
      J  = rong
      K  = Alarm Type
      L  = Time Type (Daytime / Nighttime / Overnight)
      M  = Phut Daytime
      N  = Phut Nighttime
      O  = Phut Overnight (qua ngay)
      P  = Status
      Q  = Processing Results
      R  = Gio bat dau (HH:MM)
      S  = Gio ket thuc (HH:MM)
    """
    drive_id, item_id = _resolve_excel_item()
    local_path        = _download_excel(drive_id, item_id)

    try:
        wb = openpyxl.load_workbook(local_path)
        ws = wb["Sheet1"]

        next_row = max(9, ws.max_row + 1)
        if ws.max_row >= 9:
            if ws.cell(row=ws.max_row, column=1).value is None:
                next_row = ws.max_row

        now      = datetime.datetime.utcnow() + datetime.timedelta(hours=7)
        no       = _get_next_no(ws)

        # Format ngay
        def fmt_date(d: str) -> str:
            if not d:
                return now.strftime("%d/%m/%Y")
            try:
                return datetime.datetime.strptime(d, "%Y-%m-%d").strftime("%d/%m/%Y")
            except:
                return d

        start_date_fmt = fmt_date(start_date)
        end_date_fmt   = fmt_date(end_date) if end_date else start_date_fmt

        _log(f"[EXCEL] start={start_date} {start_time} end={end_date} {end_time}")
        time_type, day_min, night_min, overnight_min = _calc_downtime(
            start_time, start_date, end_time, end_date
        )

        start_norm = _normalize_time(start_time) if start_time else ""
        end_norm   = _normalize_time(end_time)   if end_time   else ""

        def fmt_min(v):
            return f"{int(v)} phút" if v is not None else None

        ws.cell(row=next_row, column=1).value  = no              # A: STT
        ws.cell(row=next_row, column=2).value  = site_name       # B: Ten bo phan
        ws.cell(row=next_row, column=3).value  = week            # C: Week
        ws.cell(row=next_row, column=4).value  = start_date_fmt  # D: Ngay bat dau
        ws.cell(row=next_row, column=5).value  = end_date_fmt    # E: Ngay ket thuc
        ws.cell(row=next_row, column=6).value  = pic             # F: Nguoi phu trach
        # G: rong (col 7)
        ws.cell(row=next_row, column=8).value  = device          # H: Ten thiet bi
        ws.cell(row=next_row, column=9).value  = description     # I: Mo ta
        ws.cell(row=next_row, column=10).value = alarm_level     # J: Alarm Level
        ws.cell(row=next_row, column=11).value = alarm_type      # K: Alarm Type
        ws.cell(row=next_row, column=12).value = time_type       # L: Daytime/Nighttime/Overnight
        ws.cell(row=next_row, column=13).value = fmt_min(day_min)       # M: Phut Daytime
        ws.cell(row=next_row, column=14).value = fmt_min(night_min)     # N: Phut Nighttime
        ws.cell(row=next_row, column=15).value = fmt_min(overnight_min) # O: Phut Overnight
        ws.cell(row=next_row, column=16).value = status          # P: Status
        ws.cell(row=next_row, column=17).value = processing      # Q: Processing Results
        ws.cell(row=next_row, column=18).value = start_norm      # R: Gio bat dau
        ws.cell(row=next_row, column=19).value = end_norm        # S: Gio ket thuc

        wb.save(local_path)
        _log(f"[EXCEL] Saved row={next_row} no={no} | {time_type} | day={day_min} night={night_min} overnight={overnight_min}")

    finally:
        try:
            _upload_excel(drive_id, item_id, local_path)
            _log("[EXCEL] Upload thanh cong")
        finally:
            os.unlink(local_path)

    return {"row": next_row, "no": no}


def get_excel_data() -> list:
    """
    Download file Excel tu OneDrive, doc toan bo du lieu va tra ve duoi dang list of dict.
    """
    drive_id, item_id = _resolve_excel_item()
    local_path        = _download_excel(drive_id, item_id)

    try:
        wb = openpyxl.load_workbook(local_path, data_only=True)
        ws = wb["Sheet1"]
        
        data = []
        # Bat dau tu hang 9 (du lieu that)
        for row in range(9, ws.max_row + 1):
            stt = ws.cell(row=row, column=1).value
            if not stt:
                continue
                
            site_name  = ws.cell(row=row, column=2).value
            start_date = ws.cell(row=row, column=4).value
            device     = ws.cell(row=row, column=8).value
            
            # site_name, start_date, device la cac truong quan trong de thong ke
            data.append({
                "site_name":  site_name,
                "start_date": start_date,
                "device":     device,
            })
            
        return data
    finally:
        if os.path.exists(local_path):
            os.unlink(local_path)


def get_site_chart_data() -> dict:
    """
    Doc du lieu tu workbook chart moi, sheet SITE_DATA.
    Tra ve summary + danh sach site theo week de frontend ve bieu do.
    """
    local_path = _download_public_sharepoint_excel(CHART_EXCEL_SHARE_LINK)

    try:
        wb = openpyxl.load_workbook(local_path, data_only=True)
        if "SITE_DATA" not in wb.sheetnames:
            raise Exception("Khong tim thay sheet SITE_DATA")

        ws = wb["SITE_DATA"]

        header_row = None
        for row in range(1, ws.max_row + 1):
            a_val = str(ws.cell(row=row, column=1).value or "").strip().upper()
            b_val = str(ws.cell(row=row, column=2).value or "").strip().upper()
            if a_val == "SITE NAME" and b_val == "CODE SITE":
                header_row = row
                break

        start_row = (header_row + 1) if header_row else 10
        sites = []

        for row in range(start_row, ws.max_row + 1):
            site_name = ws.cell(row=row, column=1).value
            code_site = ws.cell(row=row, column=2).value

            if site_name is None and code_site is None:
                continue

            site_name_text = str(site_name or "").strip()
            code_site_text = str(code_site or "").strip()

            if not site_name_text:
                continue
            if site_name_text.upper() == "SITE NAME":
                continue

            weeks = [_to_int(ws.cell(row=row, column=col).value) for col in range(3, 8)]
            total = _to_int(ws.cell(row=row, column=8).value)
            if total == 0:
                total = sum(weeks)

            sites.append({
                "site_name": site_name_text.upper(),
                "code_site": code_site_text,
                "weeks": weeks,
                "total": total,
            })

        week_totals = [sum(site["weeks"][idx] for site in sites) for idx in range(5)]
        overall_total = sum(site["total"] for site in sites)

        return {
            "summary": {
                "week_totals": {
                    "week1": week_totals[0],
                    "week2": week_totals[1],
                    "week3": week_totals[2],
                    "week4": week_totals[3],
                    "week5": week_totals[4],
                },
                "overall_total": overall_total,
            },
            "sites": sites,
            "week_labels": ["WEEK 1", "WEEK 2", "WEEK 3", "WEEK 4", "WEEK 5"],
        }
    finally:
        if os.path.exists(local_path):
            os.unlink(local_path)